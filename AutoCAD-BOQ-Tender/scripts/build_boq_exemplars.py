"""Parse the (SOW, priced BOQ) sample pairs in archive (3) into structured
exemplars at artifacts/api-server/src/lib/boq-examples/. Each exemplar captures:

  • project metadata (ref, number, name, location, submission date, submitted to)
  • the full BOQ as a flat list of rows tagged with their 4-level hierarchy
    refs (sowRef / ourRef / subRef / srNo), exactly as the human QS wrote them
  • the SOW outline tasks the BOQ references (extracted from the SOW PDF
    table-of-contents and section headings)

Future agent prompts inject the most similar exemplar so the model learns
AIGCC house-style descriptions, units, hierarchy, and subtotal structure.
"""
import json
import os
import re
import sys
from pathlib import Path

ARCHIVE = Path(r"C:\Users\IKIO\Downloads\archive (3)")
OUT_DIR = Path(r"C:\Users\IKIO\Downloads\New\AutoCAD-BOQ-Tender\artifacts\api-server\src\lib\boq-examples")
OUT_DIR.mkdir(parents=True, exist_ok=True)


def sow_text(pdf_path: Path) -> str:
    import pdfplumber
    out = []
    with pdfplumber.open(pdf_path) as pdf:
        for p in pdf.pages:
            t = p.extract_text() or ""
            out.append(t)
    return "\n".join(out)


SECTION_RE = re.compile(r"^\s*(\d+\.\d+(?:\.\d+)?)\s+([A-Z][^\n]{2,140})$", re.MULTILINE)


def extract_sow_outline(text: str):
    """Best-effort outline: pull numbered section headings like '2.1 Foo'
    and '2.4.1 Bar'. We keep them in document order, deduplicated by ref."""
    seen = set()
    outline = []
    for m in SECTION_RE.finditer(text):
        ref = m.group(1).strip()
        title = m.group(2).strip()
        # Skip table-of-contents repeats and the page-numbered '... 12' lines.
        title = re.sub(r"\s*\.{2,}.*$", "", title).strip()
        title = re.sub(r"\s+\d+$", "", title).strip()
        if not title:
            continue
        key = (ref, title.lower())
        if key in seen:
            continue
        seen.add(key)
        outline.append({"ref": ref, "title": title})
    return outline


def coerce_ref(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        # 2.0 -> "2", 2.1 -> "2.1", 4.0 -> "4"
        if float(value).is_integer():
            return str(int(value))
        return f"{value:g}"
    return str(value).strip() or None


def parse_xlsb(path: Path, layout: str):
    """layout='1158/1159' for the 9-column SOW Ref/Our Ref/Sub Ref/Sr.No layout
    used by samples 1158 and 1159. layout='1162' for the warehouse variant
    which has an extra empty leading column.
    Returns a list of row dicts with keys: sowRef, ourRef, subRef, srNo,
    description, unit, quantity, rate, amount, remarks.
    """
    from pyxlsb import open_workbook

    rows = []
    with open_workbook(str(path)) as wb:
        sn = wb.sheets[0]
        with wb.get_sheet(sn) as sh:
            for r in sh.rows():
                vals = [c.v for c in r]
                # Trim trailing Nones
                while vals and (vals[-1] is None or vals[-1] == ""):
                    vals.pop()
                if not vals:
                    continue
                if layout == "1158":
                    # cols: SOW, Our, Sub, Sr, Description, Unit, Qty, Rate, Amount, Remarks
                    if len(vals) < 5:
                        continue
                    sow, our, sub, sr, desc = vals[:5]
                    unit = vals[5] if len(vals) > 5 else None
                    qty = vals[6] if len(vals) > 6 else None
                    rate = vals[7] if len(vals) > 7 else None
                    amount = vals[8] if len(vals) > 8 else None
                    remarks = vals[9] if len(vals) > 9 else None
                elif layout == "1159":
                    # First column is empty padding
                    if len(vals) < 6:
                        continue
                    _, sow, our, sub, sr, desc = vals[:6]
                    unit = vals[6] if len(vals) > 6 else None
                    qty = vals[7] if len(vals) > 7 else None
                    rate = vals[8] if len(vals) > 8 else None
                    amount = vals[9] if len(vals) > 9 else None
                    remarks = vals[10] if len(vals) > 10 else None
                else:
                    continue
                rows.append({
                    "sowRef": coerce_ref(sow),
                    "ourRef": coerce_ref(our),
                    "subRef": coerce_ref(sub),
                    "srNo": coerce_ref(sr),
                    "description": str(desc).strip() if desc not in (None, "") else None,
                    "unit": str(unit).strip() if unit not in (None, "") else None,
                    "quantity": qty if isinstance(qty, (int, float)) else None,
                    "rate": rate if isinstance(rate, (int, float)) else None,
                    "amount": amount if isinstance(amount, (int, float)) else None,
                    "remarks": str(remarks).strip() if remarks not in (None, "") else None,
                })
    return rows


def parse_xlsx_1162(path: Path):
    """Warehouse sample uses xlsx + a different column layout:
    cols (1-indexed): - - SOW(C) Our(D) Sub(E) Sr(F) Desc(G) Unit(H) Qty(I) Rate(J) Amount(K).
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(values_only=True):
        # row tuple already 0-indexed; pad if short
        vals = list(r)
        while len(vals) < 11:
            vals.append(None)
        sow, our, sub, sr, desc = vals[2], vals[3], vals[4], vals[5], vals[6]
        unit, qty, rate, amount = vals[7], vals[8], vals[9], vals[10]
        if all(v in (None, "") for v in (sow, our, sub, sr, desc, unit, qty, rate, amount)):
            continue
        rows.append({
            "sowRef": coerce_ref(sow),
            "ourRef": coerce_ref(our),
            "subRef": coerce_ref(sub),
            "srNo": coerce_ref(sr),
            "description": str(desc).strip() if desc not in (None, "") else None,
            "unit": str(unit).strip() if unit not in (None, "") else None,
            "quantity": qty if isinstance(qty, (int, float)) else None,
            "rate": rate if isinstance(rate, (int, float)) else None,
            "amount": amount if isinstance(amount, (int, float)) else None,
            "remarks": None,
        })
    return rows


def extract_header(rows):
    """Pull project metadata from the leading rows of a parsed BOQ.
    Looks for labels like 'Project Number', 'Project Name' etc. in the
    description column and reads the value from the same row.
    """
    meta = {}
    label_map = {
        "ref no": "refNo",
        "project number": "projectNumber",
        "project name": "projectName",
        "project location": "projectLocation",
        "submission date": "submissionDate",
        "submitted to": "submittedTo",
    }
    for r in rows[:15]:
        # The "value" lives in whichever cell holds the actual content.
        # For these samples, header rows have label in one cell and value in another.
        # We treat the description-col + srNo-col + other refs all as candidates.
        candidates = [r.get("description"), r.get("srNo"), r.get("subRef"), r.get("ourRef"), r.get("sowRef"), r.get("unit"), r.get("quantity")]
        text_cells = [str(c) for c in candidates if c not in (None, "")]
        for c in text_cells:
            low = c.lower().strip()
            for label, key in label_map.items():
                if low.startswith(label):
                    rest = c[len(label):].strip(" :")
                    if rest:
                        meta.setdefault(key, rest)
                    else:
                        # value sits in a different cell on the same row
                        for other in text_cells:
                            if other.lower().strip().startswith(label):
                                continue
                            meta.setdefault(key, other)
                            break
    # Also catch the "Ref No: AIGCC/AASAB/QO/1158/25" pattern that doesn't have a label cell.
    for r in rows[:6]:
        for c in (r.get("sowRef"), r.get("ourRef"), r.get("description")):
            if isinstance(c, str) and "Ref No" in c and "AIGCC" in c:
                meta.setdefault("refNo", c.split("Ref No:")[-1].strip())
    return meta


def is_section_header(row):
    """A row that has refs but no unit/quantity is a section/subsection header."""
    return (
        row.get("description")
        and (row.get("sowRef") or row.get("ourRef") or row.get("subRef") or row.get("srNo"))
        and row.get("unit") in (None, "")
        and row.get("quantity") in (None, "")
    )


def is_subtotal_row(row):
    d = (row.get("description") or "").lower()
    return d.startswith("total amount") or d.startswith("grand total")


def cleanup_items(rows):
    """Drop header rows (Project Name etc.), subtotal rows, and empty rows.
    Tag every remaining row with a kind: 'section' | 'item'."""
    out = []
    for r in rows:
        d = r.get("description") or ""
        # Drop the header-block labels
        if d.lower().strip() in (
            "project number", "project name", "project location", "submission date", "submitted to",
            "sow ref. no.", "our ref.no.", "our ref. no.", "sub. ref.", "sr.no.",
            "description", "unit", "quantity", "rate (in kwd)", "amount (in kwd)", "remarks",
        ):
            continue
        if isinstance(d, str) and d.lower().startswith("ref no:") and "aigcc" in d.lower():
            continue
        if is_subtotal_row(r):
            out.append({**r, "kind": "subtotal"})
            continue
        if is_section_header(r):
            out.append({**r, "kind": "section"})
            continue
        if d.strip():
            out.append({**r, "kind": "item"})
    return out


def summarize_units(items):
    counts = {}
    for it in items:
        if it.get("kind") != "item":
            continue
        u = it.get("unit")
        if not u:
            continue
        counts[u] = counts.get(u, 0) + 1
    return counts


def build_exemplar(slug, sow_pdf, boq_path, layout, sow_chapter_focus=None):
    print(f"== {slug}")
    text = sow_text(sow_pdf)
    outline = extract_sow_outline(text)
    if layout == "xlsx":
        raw = parse_xlsx_1162(boq_path)
    else:
        raw = parse_xlsb(boq_path, layout)
    meta = extract_header(raw)
    items = cleanup_items(raw)
    # Per-section breakdown: a list of {ref, title, items: [...]}
    sections = []
    current = None
    for row in items:
        if row["kind"] == "section":
            if current:
                sections.append(current)
            current = {
                "sowRef": row.get("sowRef"),
                "ourRef": row.get("ourRef"),
                "subRef": row.get("subRef"),
                "srNo": row.get("srNo"),
                "title": row.get("description"),
                "items": [],
                "subtotalLabel": None,
            }
        elif row["kind"] == "subtotal":
            if current:
                current["subtotalLabel"] = row.get("description")
                sections.append(current)
                current = None
        elif row["kind"] == "item":
            if current is None:
                current = {"sowRef": None, "ourRef": None, "subRef": None, "srNo": None,
                          "title": None, "items": [], "subtotalLabel": None}
            current["items"].append({
                "sowRef": row.get("sowRef"),
                "ourRef": row.get("ourRef"),
                "subRef": row.get("subRef"),
                "srNo": row.get("srNo"),
                "description": row.get("description"),
                "unit": row.get("unit"),
                "quantity": row.get("quantity"),
            })
    if current:
        sections.append(current)
    exemplar = {
        "slug": slug,
        "source": {
            "sowPdf": os.path.basename(sow_pdf),
            "boq": os.path.basename(boq_path),
        },
        "meta": meta,
        "units": summarize_units(items),
        "sowOutline": outline,
        "sections": sections,
        "totals": {"flatItemCount": sum(1 for r in items if r["kind"] == "item")},
    }
    return exemplar


def main():
    samples = [
        ("1158-camp-moreell-lift-station",
         ARCHIVE / "25.7.3 SOW Camp Moreell Lift Station Repair.pdf",
         ARCHIVE / "1158-Priced BOQ-Camp Morell Lift station repair-BOQ...xlsb",
         "1158"),
        ("1159-mwd-obedience-area",
         ARCHIVE / "2025.7.28 SOW 24-1020 Repair MWD Obedience Area.pdf",
         ARCHIVE / "1159-Priced BOQ-Repair MWD Obedience Area..xlsb",
         "1159"),
        ("1162-cargo-warehouse",
         None,  # SOW is .docx — we read it separately below
         ARCHIVE / "1162-Construction of Communication Cargo warehouse-Ramana Reddy-Final one - Copy.xlsx",
         "xlsx"),
    ]
    manifest = []
    for slug, sow_pdf, boq, layout in samples:
        if sow_pdf is None:
            # 1162 uses a .docx SOW; pull text via python-docx
            from docx import Document
            docx_path = ARCHIVE / "Attachment 9.2 SOW 22-1023 Comm Cargo Warehouse_25 Feb 25.docx"
            doc = Document(str(docx_path))
            text = "\n".join(p.text for p in doc.paragraphs)
            outline = extract_sow_outline(text)
            raw = parse_xlsx_1162(boq)
            meta = extract_header(raw)
            items = cleanup_items(raw)
            sections = []
            current = None
            for row in items:
                if row["kind"] == "section":
                    if current:
                        sections.append(current)
                    current = {
                        "sowRef": row.get("sowRef"),
                        "ourRef": row.get("ourRef"),
                        "subRef": row.get("subRef"),
                        "srNo": row.get("srNo"),
                        "title": row.get("description"),
                        "items": [],
                        "subtotalLabel": None,
                    }
                elif row["kind"] == "subtotal":
                    if current:
                        current["subtotalLabel"] = row.get("description")
                        sections.append(current)
                        current = None
                else:
                    if current is None:
                        current = {"sowRef": None, "ourRef": None, "subRef": None, "srNo": None,
                                   "title": None, "items": [], "subtotalLabel": None}
                    current["items"].append({
                        "sowRef": row.get("sowRef"),
                        "ourRef": row.get("ourRef"),
                        "subRef": row.get("subRef"),
                        "srNo": row.get("srNo"),
                        "description": row.get("description"),
                        "unit": row.get("unit"),
                        "quantity": row.get("quantity"),
                    })
            if current:
                sections.append(current)
            exemplar = {
                "slug": slug,
                "source": {"sowDocx": docx_path.name, "boq": boq.name},
                "meta": meta,
                "units": summarize_units(items),
                "sowOutline": outline,
                "sections": sections,
                "totals": {"flatItemCount": sum(1 for r in items if r["kind"] == "item")},
            }
        else:
            exemplar = build_exemplar(slug, sow_pdf, boq, layout)
        out_file = OUT_DIR / f"{slug}.json"
        out_file.write_text(json.dumps(exemplar, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"  wrote {out_file.relative_to(OUT_DIR.parent.parent.parent.parent)}")
        manifest.append({
            "slug": slug,
            "source": exemplar["source"],
            "projectNumber": exemplar["meta"].get("projectNumber"),
            "projectName": exemplar["meta"].get("projectName"),
            "sectionCount": len(exemplar["sections"]),
            "itemCount": exemplar["totals"]["flatItemCount"],
            "topUnits": sorted(exemplar["units"].items(), key=lambda kv: -kv[1])[:6],
        })
    (OUT_DIR / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    print("Manifest:", json.dumps(manifest, indent=2))


if __name__ == "__main__":
    main()
